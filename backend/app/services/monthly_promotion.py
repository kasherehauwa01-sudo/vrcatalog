from __future__ import annotations

import json
import logging
import smtplib
import ssl
import threading
import uuid
from datetime import datetime, timedelta
from email.message import EmailMessage
from html import escape
from io import BytesIO
from time import perf_counter

from cryptography.fernet import Fernet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import event, inspect, text
from sqlalchemy.orm import Session

from app.core.config import settings
from app.db.session import SessionLocal
from app.models.catalog import MailSetting, NotificationEmailHistory, NotificationScenarioSetting, Product, ProductPromotionState, ProductTypeChange, ProductTypeSetting
from app.services.logging import add_log

PROMOTION_VALUE = "Акция месяца"
SCENARIO_CODE = "monthly_promotion"
MOSCOW_OFFSET_HOURS = 3
logger = logging.getLogger(__name__)
_run_lock = threading.Lock()


def article_key(product: Product) -> str:
    """Возвращает стабильный ключ товара: нормализованный артикул либо код карточки."""
    return str(product.article or product.code).strip()


def normalize_month_promo(value: object) -> bool:
    """Нормализует реальные представления признака «Акция месяца»."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    normalized = " ".join(str(value or "").replace("\xa0", " ").split()).casefold()
    return normalized in {PROMOTION_VALUE.casefold(), "да", "true", "1"}


def normalized_promo_value(value: object) -> str | None:
    if normalize_month_promo(value):
        return PROMOTION_VALUE
    normalized = " ".join(str(value or "").replace("\xa0", " ").split())
    return normalized or None


def initialize_promotion_snapshot(db: Session) -> int:
    """Создаёт первичный снимок без уведомлений; повторный вызов идемпотентен."""
    if db.bind and db.bind.dialect.name == "postgresql":
        db.execute(text("SELECT pg_advisory_xact_lock(7242026)"))
    if db.query(ProductPromotionState.id).first():
        return 0
    type_names = dict(db.query(ProductTypeSetting.code, ProductTypeSetting.name).all())
    products = db.query(Product).order_by(Product.id).all()
    states: dict[str, ProductPromotionState] = {}
    duplicates = 0
    for product in products:
        key = article_key(product)
        display = type_names.get(product.product_type, product.product_type)
        if key in states:
            duplicates += 1
            if states[key].promo != normalize_month_promo(display):
                add_log(db, "promotion_snapshot_conflict", f"Conflicting duplicate article={key} promo_values=[{states[key].promo}, {normalize_month_promo(display)}]", "error")
            continue
        states[key] = ProductPromotionState(
            article_key=key, product_id=product.id, promo=normalize_month_promo(display),
            current_value=normalized_promo_value(display), version=0, updated_at=datetime.utcnow(),
        )
    db.add_all(states.values())
    add_log(db, "promotion_snapshot_initialized", f"Initial catalog snapshot created: {len(states)} products, no notifications generated; duplicates={duplicates}")
    db.commit()
    return len(states)


def initialize_product_promotion_state(db: Session, product: Product) -> None:
    """Добавляет в snapshot новую карточку без создания ложного события."""
    key = article_key(product)
    if db.query(ProductPromotionState.id).filter(ProductPromotionState.article_key == key).first():
        return
    type_name = db.query(ProductTypeSetting.name).filter(ProductTypeSetting.code == product.product_type).scalar()
    display = type_name or product.product_type
    db.add(ProductPromotionState(
        article_key=key, product_id=product.id, promo=normalize_month_promo(display),
        current_value=normalized_promo_value(display), version=0, updated_at=datetime.utcnow(),
    ))


def _fernet() -> Fernet:
    import base64
    import hashlib

    key = base64.urlsafe_b64encode(hashlib.sha256(settings.secret_key.encode()).digest())
    return Fernet(key)


def encrypt_password(password: str) -> str:
    return _fernet().encrypt(password.encode()).decode() if password else ""


def decrypt_password(value: str) -> str:
    return _fernet().decrypt(value.encode()).decode() if value else ""


def get_mail_setting(db: Session) -> MailSetting:
    value = db.query(MailSetting).order_by(MailSetting.id).first()
    if value:
        return value
    value = MailSetting(updated_at=datetime.utcnow())
    db.add(value)
    db.commit()
    db.refresh(value)
    return value


def get_scenario_setting(db: Session) -> NotificationScenarioSetting:
    value = db.query(NotificationScenarioSetting).filter_by(code=SCENARIO_CODE).first()
    if value:
        return value
    value = NotificationScenarioSetting(code=SCENARIO_CODE, updated_at=datetime.utcnow())
    db.add(value)
    db.commit()
    db.refresh(value)
    return value


def recipients(setting: NotificationScenarioSetting) -> list[str]:
    try:
        return [str(item).strip() for item in json.loads(setting.recipients_json) if str(item).strip()]
    except (TypeError, ValueError):
        return []


def _smtp(mail: MailSetting):
    context = ssl.create_default_context()
    if mail.encryption == "ssl":
        client = smtplib.SMTP_SSL(mail.smtp_host, mail.smtp_port, timeout=15, context=context)
    else:
        client = smtplib.SMTP(mail.smtp_host, mail.smtp_port, timeout=15)
        if mail.encryption == "starttls":
            client.starttls(context=context)
    if mail.username:
        client.login(mail.username, decrypt_password(mail.encrypted_password))
    return client


def check_connection(db: Session, mail: MailSetting | None = None) -> tuple[bool, str]:
    mail = mail or get_mail_setting(db)
    try:
        with _smtp(mail) as client:
            client.noop()
        mail.connection_status = "connected"
        mail.last_success_at = datetime.utcnow()
        mail.last_error = None
        result = (True, "Подключено")
    except Exception as exc:  # noqa: BLE001
        mail.connection_status = "error"
        mail.last_error = str(exc)
        result = (False, f"Ошибка подключения: {exc}")
    db.commit()
    return result


def send_email(
    db: Session,
    to: list[str],
    subject: str,
    html: str,
    attachment: bytes | None = None,
    attachment_name: str = "Изменения товаров Акция месяца.xlsx",
) -> None:
    mail = get_mail_setting(db)
    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = f"{mail.sender_name} <{mail.sender_email}>" if mail.sender_name else mail.sender_email
    message["To"] = ", ".join(to)
    message.set_content("Для просмотра письма требуется HTML-клиент.")
    message.add_alternative(html, subtype="html")
    if attachment is not None:
        message.add_attachment(
            attachment,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment_name,
        )
    with _smtp(mail) as client:
        client.send_message(message)
    mail.connection_status = "connected"
    mail.last_success_at = datetime.utcnow()
    mail.last_sent_at = datetime.utcnow()
    mail.last_error = None
    db.flush()


def _change_sections(changes: list[ProductTypeChange]):
    return (
        ("Добавлены в Акцию месяца", [item for item in changes if normalize_month_promo(item.new_value) and not normalize_month_promo(item.old_value)]),
        ("Исключены из Акции месяца", [item for item in changes if normalize_month_promo(item.old_value) and not normalize_month_promo(item.new_value)]),
    )


def build_preview(changes: list[ProductTypeChange]) -> str:
    sections = []
    for title, selected in _change_sections(changes):
        if not selected:
            continue
        rows = "".join(
            f"<tr><td>{escape(item.product_code or '')}</td><td>{escape(item.article or '')}</td><td>{escape(item.product_name)}</td>"
            f"<td>{escape(item.old_value or '')}</td><td>{escape(item.new_value or '')}</td>"
            f"<td>{item.changed_at:%d.%m.%Y %H:%M}</td></tr>"
            for item in selected
        )
        sections.append(
            f"<h2>{title}</h2><table border='1' cellpadding='6' cellspacing='0'>"
            "<tr><th>Код</th><th>Артикул</th><th>Наименование</th><th>Было</th><th>Стало</th><th>Время изменения</th></tr>"
            f"{rows}</table>"
        )
    return "".join(sections)


def build_attachment(changes: list[ProductTypeChange]) -> bytes:
    """Формирует XLSX со структурой разделов и колонок как в теле уведомления."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Акция месяца"
    headers = ("Код", "Артикул", "Наименование", "Было", "Стало", "Время изменения")
    for title, selected in _change_sections(changes):
        if not selected:
            continue
        worksheet.append([title])
        worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=len(headers))
        worksheet.cell(worksheet.max_row, 1).font = Font(bold=True, size=14)
        worksheet.append(headers)
        for cell in worksheet[worksheet.max_row]:
            cell.font = Font(bold=True)
        for item in selected:
            worksheet.append((
                item.product_code or "", item.article or "", item.product_name,
                item.old_value or "", item.new_value or "", item.changed_at,
            ))
            worksheet.cell(worksheet.max_row, len(headers)).number_format = "DD.MM.YYYY HH:MM"
        worksheet.append([])
    for column, width in zip("ABCDEF", (18, 18, 60, 24, 24, 21), strict=True):
        worksheet.column_dimensions[column].width = width
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def consolidate_changes(changes: list[ProductTypeChange]) -> list[ProductTypeChange]:
    """Сворачивает старые/конкурентные события до одного итогового перехода на артикул."""
    grouped: dict[str, list[ProductTypeChange]] = {}
    for item in changes:
        key = str(item.article or item.product_id).strip()
        grouped.setdefault(key, []).append(item)
    result: list[ProductTypeChange] = []
    for items in grouped.values():
        items.sort(key=lambda item: (item.changed_at, item.id or 0))
        initial = normalize_month_promo(items[0].old_value)
        final = normalize_month_promo(items[-1].new_value)
        if initial == final:
            continue
        wanted = final
        representative = next(
            item for item in reversed(items)
            if normalize_month_promo(item.new_value) == wanted
            and normalize_month_promo(item.old_value) != wanted
        )
        result.append(representative)
    return sorted(result, key=lambda item: (item.changed_at, item.id or 0))


def log_scenario_run(db: Session, result: dict, started: float) -> None:
    add_log(db, "monthly_promotion_scenario", json.dumps({
        **{key: value for key, value in result.items() if key != "html"},
        "duration_ms": round((perf_counter() - started) * 1000, 3),
    }, ensure_ascii=False), "error" if result.get("status") == "error" else "info")
    db.commit()


def run_scenario(db: Session, *, force: bool = False) -> dict:
    started = perf_counter()
    scenario = get_scenario_setting(db)
    targets = recipients(scenario)
    pending_changes = (
        db.query(ProductTypeChange).filter(
            ProductTypeChange.processed.is_(False), ProductTypeChange.claim_token.is_(None),
        )
        .order_by(ProductTypeChange.changed_at).with_for_update(skip_locked=True).all()
    )
    changes = consolidate_changes(pending_changes)
    result = {"changes": len(changes), "pending_rows": len(pending_changes), "sent": 0, "recipients": targets, "html": build_preview(changes)}
    if not force and not scenario.enabled:
        result["status"] = "disabled"
        log_scenario_run(db, result, started)
        return result
    if not changes:
        result["status"] = "empty"
        now = datetime.utcnow()
        for item in pending_changes:
            item.processed = True
            item.processed_at = now
        log_scenario_run(db, result, started)
        return result
    if not targets:
        result["status"] = "no_recipients"
        log_scenario_run(db, result, started)
        return result
    subject = 'Изменения товаров "Акция месяца"'
    claim_token = uuid.uuid4().hex
    claimed_ids = [item.id for item in pending_changes]
    for item in pending_changes:
        item.claim_token = claim_token
    db.commit()
    smtp_completed = False
    try:
        send_email(db, targets, subject, result["html"], build_attachment(changes))
        smtp_completed = True
        now = datetime.utcnow()
        claimed_changes = db.query(ProductTypeChange).filter(ProductTypeChange.id.in_(claimed_ids), ProductTypeChange.claim_token == claim_token).all()
        for item in claimed_changes:
            item.processed = True
            item.processed_at = now
            item.claim_token = None
        result.update(status="sent", sent=1)
        db.add(NotificationEmailHistory(
            scenario_code=SCENARIO_CODE, sent_at=now, recipients_json=json.dumps(targets, ensure_ascii=False),
            subject=subject, body_html=result["html"], status="sent",
            duration_ms=round((perf_counter() - started) * 1000, 3),
        ))
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        if not smtp_completed:
            db.query(ProductTypeChange).filter(
                ProductTypeChange.id.in_(claimed_ids), ProductTypeChange.claim_token == claim_token,
            ).update({ProductTypeChange.claim_token: None}, synchronize_session=False)
        mail = get_mail_setting(db)
        mail.connection_status = "error"
        mail.last_error = str(exc)
        result.update(status="error", error=str(exc))
        db.add(NotificationEmailHistory(
            scenario_code=SCENARIO_CODE, sent_at=datetime.utcnow(), recipients_json=json.dumps(targets, ensure_ascii=False),
            subject=subject, body_html=result["html"], status="error", error_message=str(exc),
            duration_ms=round((perf_counter() - started) * 1000, 3),
        ))
    finally:
        log_scenario_run(db, result, started)
    return result


def run_scheduled_if_due() -> None:
    if not _run_lock.acquire(blocking=False):
        return
    db = SessionLocal()
    try:
        scenario = get_scenario_setting(db)
        now = datetime.utcnow() + timedelta(hours=MOSCOW_OFFSET_HOURS)
        current = now.strftime("%H:%M")
        today = now.strftime("%Y-%m-%d")
        if scenario.enabled and current >= scenario.send_time and scenario.last_run_date != today:
            run_scenario(db)
            scenario = get_scenario_setting(db)
            scenario.last_run_date = today
            db.commit()
    finally:
        db.close()
        _run_lock.release()


@event.listens_for(Session, "before_flush")
def track_product_type_changes(session: Session, _flush_context, _instances) -> None:
    source = str(session.info.get("change_source", "api"))
    dirty_products = [item for item in session.dirty if isinstance(item, Product)]
    if not dirty_products:
        return
    type_names = dict(session.query(ProductTypeSetting.code, ProductTypeSetting.name).all())
    seen = session.info.setdefault("promotion_event_keys", set())
    for product in dirty_products:
        history = inspect(product).attrs.product_type.history
        if not history.has_changes():
            continue
        new = history.added[0] if history.added else product.product_type
        new_display = type_names.get(new, new)
        key = article_key(product)
        owner = session.info.get("promotion_article_owner", {}).get(key)
        if owner and owner != product.code:
            add_log(session, "promotion_duplicate_product_skipped", f"Skipped duplicate product article={key}; canonical_code={owner}; skipped_code={product.code}")
            continue
        state = (
            session.query(ProductPromotionState)
            .filter(ProductPromotionState.article_key == key)
            .with_for_update()
            .first()
        )
        new_promo = normalize_month_promo(new_display)
        new_value = normalized_promo_value(new_display)
        if state is None:
            session.add(ProductPromotionState(
                article_key=key, product_id=product.id, promo=new_promo,
                current_value=new_value, version=0, updated_at=datetime.utcnow(),
            ))
            add_log(session, "promotion_snapshot_product_initialized", f"Initial product snapshot created: article={key}, no notification generated")
            continue
        if state.promo == new_promo:
            state.product_id = product.id
            state.current_value = new_value
            continue
        transition_key = (key, state.promo, new_promo, state.version + 1)
        if transition_key in seen:
            add_log(session, "promotion_duplicate_event_skipped", f"Skipped duplicate promo event: article={key}")
            continue
        seen.add(transition_key)
        old_value = state.current_value
        state.promo = new_promo
        state.current_value = new_value
        state.product_id = product.id
        state.version += 1
        state.updated_at = datetime.utcnow()
        event_key = f"{key}:{state.version}:{int(not new_promo)}->{int(new_promo)}"
        add_log(session, "promotion_change_detected", f"PROMO CHANGE article={key} old={not new_promo} new={new_promo}")
        session.add(ProductTypeChange(
            product_id=product.id,
            product_code=product.code,
            article=product.article,
            product_name=product.name,
            old_value=old_value,
            new_value=new_value,
            source=source,
            changed_at=datetime.utcnow(),
            event_key=event_key,
        ))
