import csv
import json
import tempfile
from copy import copy
from concurrent.futures import ThreadPoolExecutor, wait
from io import StringIO, BytesIO
from pathlib import Path
from urllib.parse import quote, urlsplit, urlunsplit
from urllib.request import Request, urlopen

from typing import Annotated, Callable, Literal

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PillowImage, UnidentifiedImageError
from sqlalchemy.orm import Session, selectinload

from app.db.session import get_db
from app.importer.xml_importer import XMLCatalogImporter
from app.models.catalog import Favorite, Notification, NotificationEmailHistory, Product, ProductTypeSetting, ServiceLog, Stock, ViewHistory, WarehouseSetting
from app.schemas.catalog import AutoImportStateOut, FtpConnectionTestOut, MailSettingIn, MailSettingOut, MetaOut, NotificationHistoryOut, NotificationOut, ProductDetailOut, ProductListOut, ProductPageOut, ProductTypeUpdateIn, ScenarioRunOut, ScenarioSettingIn, ScenarioSettingOut, ScenarioSummaryOut, ServiceLogOut, TestMailIn, WarehouseSettingIn, WarehouseSettingOut, ProductTypeSettingIn, ProductTypeSettingOut, XmlServerSettingIn, XmlServerSettingOut
from app.services.catalog import decorate, list_filters, meta, product_query, paginated_products
from app.services.logging import add_log
from app.services.xml_auto_import import get_auto_import_state, get_xml_server_setting, start_manual_import, test_connection
from app.services.monthly_promotion import check_connection as check_mail_connection, encrypt_password, get_mail_setting, get_scenario_setting, recipients as scenario_recipients, run_scenario, send_email

router = APIRouter()

@router.get("/health")
def health():
    return {"status": "ok"}

@router.post("/import", response_model=MetaOut)
def upload_xml(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith(".xml"):
        raise HTTPException(400, "Загрузите XML-файл")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xml") as tmp:
        tmp.write(file.file.read())
        path = Path(tmp.name)
    try:
        XMLCatalogImporter().import_file(db, path, file.filename)
    except Exception as exc:
        path.unlink(missing_ok=True)
        raise HTTPException(400, f"Ошибка импорта XML. Файл: {file.filename}. Причина: {exc}") from exc
    path.unlink(missing_ok=True)
    return meta(db)

@router.get("/products", response_model=list[ProductListOut])
def products(db: Session = Depends(get_db), limit: int = 60, offset: int = 0, search: str | None = None, section: str | None = None, manufacturer: str | None = None, brand: str | None = None, manager: str | None = None, country: str | None = None, material: str | None = None, color: str | None = None, in_stock: str | None = None, price_min: str | None = None, price_max: str | None = None, stock_min: str | None = None, stock_max: str | None = None, warehouse: str | None = None, product_type: str | None = None, only_new: Annotated[bool, Query(alias="onlyNew")] = False):
    params = locals(); params.pop("db"); params.pop("limit"); params.pop("offset")
    type_names = {item.code: item.name for item in db.query(ProductTypeSetting).all()}
    return [decorate(p, type_names) for p in product_query(db, params).offset(offset).limit(limit).all()]


@router.get("/products/count")
def products_count(db: Session = Depends(get_db), search: str | None = None, section: str | None = None, manufacturer: str | None = None, brand: str | None = None, manager: str | None = None, country: str | None = None, material: str | None = None, color: str | None = None, in_stock: str | None = None, price_min: str | None = None, price_max: str | None = None, stock_min: str | None = None, stock_max: str | None = None, warehouse: str | None = None, product_type: str | None = None, only_new: Annotated[bool, Query(alias="onlyNew")] = False):
    params = locals(); params.pop("db")
    return {"count": product_query(db, params).count()}


@router.get("/products/search", response_model=ProductPageOut)
def search_products(
    db: Session = Depends(get_db),
    search: Annotated[str | None, Query(max_length=255)] = None,
    id: Annotated[int | None, Query(ge=1)] = None,
    name: Annotated[str | None, Query(max_length=512)] = None,
    code: Annotated[str | None, Query(max_length=2000)] = None,
    article: Annotated[str | None, Query(max_length=2000)] = None,
    barcode: Annotated[str | None, Query(max_length=2000)] = None,
    section: Annotated[str | None, Query(max_length=2000)] = None,
    manufacturer: Annotated[str | None, Query(max_length=2000)] = None,
    brand: Annotated[str | None, Query(max_length=2000)] = None,
    manager: Annotated[str | None, Query(max_length=2000)] = None,
    country: Annotated[str | None, Query(max_length=2000)] = None,
    material: Annotated[str | None, Query(max_length=2000)] = None,
    color: Annotated[str | None, Query(max_length=2000)] = None,
    product_type: Annotated[str | None, Query(alias="productType", max_length=2000)] = None,
    warehouse: Annotated[str | None, Query(max_length=2000)] = None,
    availability: Literal["all", "in_stock", "out_of_stock"] = "all",
    in_stock_only: Annotated[bool, Query(alias="inStockOnly")] = True,
    exclude_yyy: Annotated[bool, Query(alias="excludeYyy")] = True,
    only_new: Annotated[bool, Query(alias="onlyNew")] = False,
    quantity_from: Annotated[float | None, Query(alias="quantityFrom")] = None,
    quantity_to: Annotated[float | None, Query(alias="quantityTo")] = None,
    price_from: Annotated[float | None, Query(alias="priceFrom", ge=0)] = None,
    price_to: Annotated[float | None, Query(alias="priceTo", ge=0)] = None,
    property: Annotated[list[str] | None, Query()] = None,
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(alias="pageSize")] = 100,
    sort: Literal["updated_at", "is_new", "id", "name", "article", "code", "price", "quantity"] = "updated_at",
    order: Literal["asc", "desc"] | None = None,
):
    if page_size not in {20, 50, 100}:
        raise HTTPException(422, "pageSize должен быть равен 20, 50 или 100")
    if quantity_from is not None and quantity_to is not None and quantity_from > quantity_to:
        raise HTTPException(422, "Минимальное количество не может быть больше максимального")
    if price_from is not None and price_to is not None and price_from > price_to:
        raise HTTPException(422, "Минимальная цена не может быть больше максимальной")
    properties: dict[str, list[str]] = {}
    for item in property or []:
        property_name, separator, value = item.partition(":")
        if not separator or not property_name.strip() or not value.strip():
            raise HTTPException(422, "Свойство должно иметь формат «Название:Значение»")
        properties.setdefault(property_name.strip(), []).append(value.strip())
    params = {
        "search": search,
        "id": id,
        "name": name,
        "code": code,
        "article": article,
        "barcode": barcode,
        "section": section,
        "manufacturer": manufacturer,
        "brand": brand,
        "manager": manager,
        "country": country,
        "material": material,
        "color": color,
        "product_type": product_type,
        "warehouse": warehouse,
        "availability": availability,
        "in_stock_only": in_stock_only,
        "exclude_yyy": exclude_yyy,
        "only_new": only_new,
        "quantity_from": quantity_from,
        "quantity_to": quantity_to,
        "price_from": price_from,
        "price_to": price_to,
        "properties": properties,
        "page": page,
        "page_size": page_size,
        "sort": sort,
        "order": order,
    }
    items, pagination = paginated_products(db, params)
    type_names = {item.code: item.name for item in db.query(ProductTypeSetting).all()}
    return {"items": [decorate(item, type_names) for item in items], "pagination": pagination}

@router.delete("/products")
def delete_products(product_ids: list[int] = Body(...), db: Session = Depends(get_db)):
    deleted = db.query(Product).filter(Product.id.in_(product_ids)).delete(synchronize_session=False)
    add_log(db, "products_delete", f"Удалено товаров: {deleted}")
    db.commit()
    return {"deleted": deleted}

@router.get("/products/{product_id}", response_model=ProductDetailOut)
def product_detail(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).options(selectinload(Product.prices), selectinload(Product.stocks), selectinload(Product.properties), selectinload(Product.analogs), selectinload(Product.barcodes), selectinload(Product.images)).get(product_id)
    if not product:
        raise HTTPException(404, "Товар не найден")
    db.add(ViewHistory(product_id=product_id)); db.commit()
    type_names = {item.code: item.name for item in db.query(ProductTypeSetting).all()}
    warehouse_names = {item.code: item.name for item in db.query(WarehouseSetting).all()}
    for stock in product.stocks:
        stock.warehouse_name = warehouse_names.get(stock.warehouse, stock.warehouse)
    return decorate(product, type_names)


@router.patch("/products/{product_id}/product-type")
def update_product_product_type(product_id: int, payload: ProductTypeUpdateIn, db: Session = Depends(get_db)):
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(404, "Товар не найден")
    db.info["change_source"] = "manual"
    product.product_type = payload.product_type.strip() if payload.product_type else None
    db.commit()
    return {"ok": True, "product_type": product.product_type}


@router.get("/xml-server-settings", response_model=XmlServerSettingOut)
def xml_server_settings(db: Session = Depends(get_db)):
    return get_xml_server_setting(db)

@router.put("/xml-server-settings", response_model=XmlServerSettingOut)
def update_xml_server_settings(payload: XmlServerSettingIn, db: Session = Depends(get_db)):
    if payload.protocol.upper() != "FTP":
        raise HTTPException(400, "Пока поддерживается только FTP")
    setting = get_xml_server_setting(db)
    setting.protocol = payload.protocol.upper()
    setting.host = payload.host.strip()
    setting.port = payload.port
    setting.username = payload.username.strip()
    setting.password = payload.password
    setting.xml_dir = payload.xml_dir.strip() or "/"
    setting.connection_attempts = payload.connection_attempts
    setting.retry_delay_seconds = payload.retry_delay_seconds
    db.commit()
    db.refresh(setting)
    return setting

@router.post("/xml-server-settings/test", response_model=FtpConnectionTestOut)
def test_xml_server_settings(db: Session = Depends(get_db)):
    success, message = test_connection(db)
    return {"success": success, "message": message}


def mail_setting_response(item) -> dict:
    return {
        "smtp_host": item.smtp_host,
        "smtp_port": item.smtp_port,
        "encryption": item.encryption,
        "username": item.username,
        "password_configured": bool(item.encrypted_password),
        "sender_name": item.sender_name,
        "sender_email": item.sender_email,
        "connection_status": item.connection_status,
        "last_success_at": item.last_success_at,
        "last_sent_at": item.last_sent_at,
        "last_error": item.last_error,
    }


@router.get("/mail-settings", response_model=MailSettingOut)
def mail_settings(db: Session = Depends(get_db)):
    return mail_setting_response(get_mail_setting(db))


@router.put("/mail-settings", response_model=MailSettingOut)
def update_mail_settings(payload: MailSettingIn, db: Session = Depends(get_db)):
    item = get_mail_setting(db)
    for field in ("smtp_host", "smtp_port", "encryption", "username", "sender_name", "sender_email"):
        setattr(item, field, getattr(payload, field))
    if payload.password:
        item.encrypted_password = encrypt_password(payload.password)
    db.commit()
    check_mail_connection(db, item)
    return mail_setting_response(item)


@router.post("/mail-settings/test")
def send_test_mail(payload: TestMailIn, db: Session = Depends(get_db)):
    try:
        send_email(db, [payload.email], "Тест уведомлений VR Catalog", "<h2>Тестовое письмо отправлено успешно</h2>")
        db.commit()
        return {"success": True, "message": "Тестовое письмо отправлено успешно."}
    except Exception as exc:
        db.rollback()
        item = get_mail_setting(db)
        item.connection_status = "error"
        item.last_error = str(exc)
        db.commit()
        return {"success": False, "message": f"Ошибка отправки: {exc}"}


@router.get("/notification-scenarios/monthly-promotion", response_model=ScenarioSettingOut)
def monthly_promotion_settings(db: Session = Depends(get_db)):
    item = get_scenario_setting(db)
    return {"code": item.code, "enabled": item.enabled, "send_time": item.send_time, "recipients": scenario_recipients(item)}


@router.get("/notification-scenarios", response_model=list[ScenarioSummaryOut])
def notification_scenarios(db: Session = Depends(get_db)):
    item = get_scenario_setting(db)
    return [{"code": item.code, "name": "Акция месяца", "enabled": item.enabled}]


@router.patch("/notification-scenarios/{code}/enabled", response_model=ScenarioSummaryOut)
def toggle_notification_scenario(code: str, enabled: bool = Body(embed=True), db: Session = Depends(get_db)):
    if code != "monthly_promotion":
        raise HTTPException(404, "Сценарий не найден")
    item = get_scenario_setting(db)
    item.enabled = enabled
    add_log(db, "notification_scenario_toggled", json.dumps({"scenario": code, "enabled": enabled}, ensure_ascii=False))
    db.commit()
    return {"code": item.code, "name": "Акция месяца", "enabled": item.enabled}


@router.put("/notification-scenarios/monthly-promotion", response_model=ScenarioSettingOut)
def update_monthly_promotion_settings(payload: ScenarioSettingIn, db: Session = Depends(get_db)):
    item = get_scenario_setting(db)
    item.enabled = payload.enabled
    item.send_time = payload.send_time
    item.recipients_json = json.dumps(list(dict.fromkeys(payload.recipients)), ensure_ascii=False)
    add_log(db, "notification_scenario_updated", json.dumps({"scenario": item.code, "send_time": item.send_time, "recipients": len(payload.recipients)}, ensure_ascii=False))
    db.commit()
    return {"code": item.code, "enabled": item.enabled, "send_time": item.send_time, "recipients": scenario_recipients(item)}


@router.post("/notification-scenarios/monthly-promotion/run", response_model=ScenarioRunOut)
def run_monthly_promotion(db: Session = Depends(get_db)):
    return run_scenario(db)


@router.get("/notification-scenarios/monthly-promotion/preview", response_model=ScenarioRunOut)
def preview_monthly_promotion(db: Session = Depends(get_db)):
    from app.models.catalog import ProductTypeChange
    from app.services.monthly_promotion import build_preview, consolidate_changes
    changes = consolidate_changes(db.query(ProductTypeChange).filter(ProductTypeChange.processed.is_(False)).order_by(ProductTypeChange.changed_at).all())
    html = build_preview(changes)
    add_log(db, "notification_scenario_preview", json.dumps({"scenario": "monthly_promotion", "changes": len(changes)}, ensure_ascii=False))
    db.commit()
    return {"status": "preview", "changes": len(changes), "sent": 0, "recipients": scenario_recipients(get_scenario_setting(db)), "html": html}


@router.get("/notification-scenarios/{code}/history", response_model=list[NotificationHistoryOut])
def notification_scenario_history(
    code: str,
    search: str = "",
    status: Literal["all", "sent", "error"] = "all",
    db: Session = Depends(get_db),
):
    query = db.query(NotificationEmailHistory).filter(NotificationEmailHistory.scenario_code == code)
    if search.strip():
        query = query.filter(NotificationEmailHistory.recipients_json.ilike(f"%{search.strip()}%"))
    if status != "all":
        query = query.filter(NotificationEmailHistory.status == status)
    rows = query.order_by(NotificationEmailHistory.sent_at.desc(), NotificationEmailHistory.id.desc()).limit(500).all()
    result = []
    for item in rows:
        result.append({
            "id": item.id, "scenario_code": item.scenario_code, "sent_at": item.sent_at,
            "recipients": json.loads(item.recipients_json), "subject": item.subject,
            "body_html": item.body_html, "status": item.status, "error_message": item.error_message,
            "duration_ms": item.duration_ms,
        })
    return result

@router.get("/auto-import-state", response_model=AutoImportStateOut)
def auto_import_state(db: Session = Depends(get_db)):
    return get_auto_import_state(db)

@router.post("/auto-import/run")
def run_auto_import_now():
    started = start_manual_import()
    return {"started": started}

@router.get("/filters")
def filters(
    db: Session = Depends(get_db),
    brand: str | None = None,
    manager: str | None = None,
    manufacturer: str | None = None,
    country: str | None = None,
    product_type: str | None = Query(None, alias="productType"),
    warehouse: str | None = None,
    barcode: str | None = None,
    in_stock_only: bool = Query(True, alias="inStockOnly"),
    exclude_yyy: bool = Query(True, alias="excludeYyy"),
    property: list[str] | None = Query(None),
):
    properties: dict[str, list[str]] = {}
    for item in property or []:
        property_name, separator, value = item.partition(":")
        if separator and property_name.strip() and value.strip():
            properties.setdefault(property_name.strip(), []).append(value.strip())
    return list_filters(db, {
        "brand": brand,
        "manager": manager,
        "manufacturer": manufacturer,
        "country": country,
        "product_type": product_type,
        "warehouse": warehouse,
        "barcode": barcode,
        "in_stock_only": in_stock_only,
        "exclude_yyy": exclude_yyy,
        "properties": properties,
    })

@router.get("/meta", response_model=MetaOut)
def get_meta(db: Session = Depends(get_db)):
    return meta(db)

@router.post("/favorites/{product_id}")
def toggle_favorite(product_id: int, db: Session = Depends(get_db)):
    favorite = db.get(Favorite, product_id)
    if favorite: db.delete(favorite); active = False
    else: db.add(Favorite(product_id=product_id)); active = True
    db.commit(); return {"favorite": active}


@router.get("/warehouses", response_model=list[WarehouseSettingOut])
def warehouses(db: Session = Depends(get_db)):
    return db.query(WarehouseSetting).order_by(WarehouseSetting.code).all()

@router.get("/warehouses/codes")
def warehouse_codes(db: Session = Depends(get_db)):
    codes = [code for code, in db.query(Stock.warehouse).filter(Stock.warehouse.isnot(None)).distinct().order_by(Stock.warehouse).all()]
    return {"codes": codes}

@router.post("/warehouses", response_model=WarehouseSettingOut)
def create_warehouse(payload: WarehouseSettingIn, db: Session = Depends(get_db)):
    code = payload.code.strip()
    name = payload.name.strip()
    if not code or not name:
        raise HTTPException(400, "Заполните код и имя склада")
    if db.query(WarehouseSetting).filter(WarehouseSetting.code == code).first():
        raise HTTPException(400, "Склад с таким кодом уже добавлен")
    warehouse = WarehouseSetting(code=code, name=name)
    db.add(warehouse)
    db.commit()
    db.refresh(warehouse)
    return warehouse

@router.put("/warehouses/{warehouse_id}", response_model=WarehouseSettingOut)
def update_warehouse(warehouse_id: int, payload: WarehouseSettingIn, db: Session = Depends(get_db)):
    warehouse = db.get(WarehouseSetting, warehouse_id)
    if not warehouse:
        raise HTTPException(404, "Склад не найден")
    code = payload.code.strip()
    name = payload.name.strip()
    if not code or not name:
        raise HTTPException(400, "Заполните код и имя склада")
    duplicate = db.query(WarehouseSetting).filter(WarehouseSetting.code == code, WarehouseSetting.id != warehouse_id).first()
    if duplicate:
        raise HTTPException(400, "Склад с таким кодом уже добавлен")
    warehouse.code = code
    warehouse.name = name
    db.commit()
    db.refresh(warehouse)
    return warehouse

@router.delete("/warehouses/{warehouse_id}")
def delete_warehouse(warehouse_id: int, db: Session = Depends(get_db)):
    warehouse = db.get(WarehouseSetting, warehouse_id)
    if not warehouse:
        raise HTTPException(404, "Склад не найден")
    db.delete(warehouse)
    db.commit()
    return {"deleted": True}


@router.get("/product-types", response_model=list[ProductTypeSettingOut])
def product_types(db: Session = Depends(get_db)):
    return db.query(ProductTypeSetting).order_by(ProductTypeSetting.code).all()

@router.get("/product-types/codes")
def product_type_codes(db: Session = Depends(get_db)):
    codes = [code for code, in db.query(Product.product_type).filter(Product.product_type.isnot(None)).distinct().order_by(Product.product_type).all()]
    return {"codes": codes}

@router.post("/product-types", response_model=ProductTypeSettingOut)
def create_product_type(payload: ProductTypeSettingIn, db: Session = Depends(get_db)):
    code = payload.code.strip()
    name = payload.name.strip()
    if not code or not name:
        raise HTTPException(400, "Заполните код и наименование вида товара")
    if db.query(ProductTypeSetting).filter(ProductTypeSetting.code == code).first():
        raise HTTPException(400, "Вид товара с таким кодом уже добавлен")
    item = ProductTypeSetting(code=code, name=name)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item

@router.put("/product-types/{product_type_id}", response_model=ProductTypeSettingOut)
def update_product_type(product_type_id: int, payload: ProductTypeSettingIn, db: Session = Depends(get_db)):
    item = db.get(ProductTypeSetting, product_type_id)
    if not item:
        raise HTTPException(404, "Вид товара не найден")
    code = payload.code.strip()
    name = payload.name.strip()
    if not code or not name:
        raise HTTPException(400, "Заполните код и наименование вида товара")
    duplicate = db.query(ProductTypeSetting).filter(ProductTypeSetting.code == code, ProductTypeSetting.id != product_type_id).first()
    if duplicate:
        raise HTTPException(400, "Вид товара с таким кодом уже добавлен")
    item.code = code
    item.name = name
    db.commit()
    db.refresh(item)
    return item

@router.delete("/product-types/{product_type_id}")
def delete_product_type(product_type_id: int, db: Session = Depends(get_db)):
    item = db.get(ProductTypeSetting, product_type_id)
    if not item:
        raise HTTPException(404, "Вид товара не найден")
    db.delete(item)
    db.commit()
    return {"deleted": True}


def error_notifications_query(db: Session):
    return db.query(Notification).filter(Notification.type.ilike("%error%"))

@router.get("/notifications", response_model=list[NotificationOut])
def notifications(db: Session = Depends(get_db), limit: int = 200):
    return error_notifications_query(db).order_by(Notification.created_at.desc()).limit(limit).all()

@router.get("/notifications/unread-count")
def notifications_unread_count(db: Session = Depends(get_db)):
    return {"count": error_notifications_query(db).filter(Notification.is_read.is_(False)).count()}

@router.post("/notifications/read-all")
def notifications_read_all(db: Session = Depends(get_db)):
    updated = error_notifications_query(db).filter(Notification.is_read.is_(False)).update({Notification.is_read: True}, synchronize_session=False)
    db.commit()
    return {"updated": updated}

@router.post("/notifications/{notification_id}/read")
def notification_read(notification_id: int, db: Session = Depends(get_db)):
    notification = db.get(Notification, notification_id)
    if not notification:
        raise HTTPException(404, "Уведомление не найдено")
    notification.is_read = True
    db.commit()
    return {"ok": True}

@router.get("/logs", response_model=list[ServiceLogOut])
def logs(db: Session = Depends(get_db)):
    return (
        db.query(ServiceLog)
        .filter(ServiceLog.level == "error")
        .order_by(ServiceLog.created_at.desc(), ServiceLog.id.desc())
        .limit(100)
        .all()
    )

@router.get("/export.csv")
def export_csv(db: Session = Depends(get_db), search: str | None = None):
    add_log(db, "export_csv", f"Экспорт CSV; поиск: {search or ''}")
    db.commit()
    output = StringIO(); writer = csv.writer(output); writer.writerow(["Код", "Артикул", "Название", "Раздел", "Остаток"])
    for p in product_query(db, {"search": search}).all(): writer.writerow([p.code, p.article, p.name, p.section, p.quantity])
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=products.csv"})

@router.get("/export.xlsx")
def export_xlsx(db: Session = Depends(get_db), search: str | None = None, section: str | None = None, manufacturer: str | None = None, brand: str | None = None, manager: str | None = None, country: str | None = None, material: str | None = None, color: str | None = None, in_stock: str | None = None, price_min: str | None = None, price_max: str | None = None, stock_min: str | None = None, stock_max: str | None = None, warehouse: str | None = None, product_type: str | None = None, exclude_yyy: bool = Query(True, alias="excludeYyy"), column: Annotated[list[str] | None, Query()] = None):
    params = locals(); params.pop("db"); columns = params.pop("column")
    add_log(db, "export_xlsx", f"Экспорт Excel; поиск: {search or ''}")
    db.commit()
    wb = build_export_workbook(db, params, columns)
    stream = BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=products.xlsx"})


EXPORT_MAIN_COLUMNS = {
    "code": "Код",
    "article": "Артикул",
    "photo": "Фото",
    "name": "Наименование",
    "section": "Раздел",
    "product_type": "Вид товара",
    "manufacturer": "Производитель",
    "manager": "Менеджер",
    "marking_code": "Код маркировки",
    "material": "Материал",
    "certificate": "Сертификат",
    "barcodes": "Штрихкоды",
}
LEGACY_EXPORT_COLUMNS = ["code", "article", "name", "section", "quantity"]
EXPORT_LEADING_COLUMNS = ("photo", "article", "name", "section", "code")
EXPORT_FIXED_WIDTHS = {
    "name": 50,
    "section": 12,
    "manufacturer": 20,
    "manager": 12,
    "material": 12,
    "barcodes": 17,
}
EXPORT_VALUE_WIDTH_COLUMNS = {"code", "certificate", "product_type"}


def normalize_export_property_key(value: str | None) -> str:
    """Нормализует название свойства для устойчивого поиска значений при экспорте."""
    return "".join(character for character in (value or "").casefold() if character.isalnum())


def normalize_image_url(url: str) -> str:
    """Кодирует пробелы и кириллицу в путях изображений из XML."""
    parts = urlsplit(url.strip())
    if parts.scheme not in {"http", "https"} or not parts.netloc:
        raise ValueError("Некорректный URL изображения")
    return urlunsplit((
        parts.scheme,
        parts.netloc.encode("idna").decode("ascii"),
        quote(parts.path, safe="/%:@"),
        quote(parts.query, safe="=&%:@/?"),
        "",
    ))


def download_export_image(url: str) -> BytesIO | None:
    """Загружает изображение для Excel с ограничением времени и объёма ответа."""
    try:
        normalized_url = normalize_image_url(url)
        request = Request(normalized_url, headers={
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
            "Referer": "https://volgorost.ru/",
            "User-Agent": "Mozilla/5.0 (compatible; VRCatalog Excel Export)",
        })
        with urlopen(request, timeout=3) as response:
            content = response.read(10 * 1024 * 1024 + 1)
        if len(content) > 10 * 1024 * 1024:
            return None
        source = BytesIO(content)
        with PillowImage.open(source) as image:
            image.thumbnail((100, 100))
            prepared = BytesIO()
            image.convert("RGBA" if image.mode == "RGBA" else "RGB").save(prepared, format="PNG")
            prepared.seek(0)
            return prepared
    except (OSError, ValueError, UnidentifiedImageError):
        return None


def download_export_images(
    urls: list[str],
    image_loader: Callable[[str], BytesIO | None],
) -> dict[str, bytes]:
    """Параллельно загружает уникальные фото, не задерживая экспорт дольше 15 секунд."""
    unique_urls = list(dict.fromkeys(urls))
    if not unique_urls:
        return {}
    executor = ThreadPoolExecutor(max_workers=min(24, len(unique_urls)))
    futures = {executor.submit(image_loader, url): url for url in unique_urls}
    completed, pending = wait(futures, timeout=15)
    images: dict[str, bytes] = {}
    for future in completed:
        try:
            stream = future.result()
        except Exception:
            stream = None
        if stream:
            images[futures[future]] = stream.getvalue()
    for future in pending:
        future.cancel()
    executor.shutdown(wait=False, cancel_futures=True)
    return images


def build_export_workbook(
    db: Session,
    params: dict,
    columns: list[str] | None,
    image_loader: Callable[[str], BytesIO | None] = download_export_image,
) -> Workbook:
    selected_columns = columns or LEGACY_EXPORT_COLUMNS
    warehouse_names = {item.code: item.name for item in db.query(WarehouseSetting).all()}
    product_type_names = {item.code: item.name for item in db.query(ProductTypeSetting).all()}
    allowed_columns = set(EXPORT_MAIN_COLUMNS) | {"quantity"}
    allowed_columns.update(f"price:{name}" for name in ("ЦенаОптовая", "ЦенаКорпоративная", "ЦенаРозничная"))
    allowed_columns.update(f"stock:{code}" for code in warehouse_names)
    unknown_columns = set(selected_columns) - allowed_columns
    if unknown_columns:
        raise HTTPException(422, f"Неизвестные колонки экспорта: {', '.join(sorted(unknown_columns))}")

    # Основные поля сохраняют одинаковый порядок независимо от порядка выбора пользователя.
    selected_columns = [
        *(column for column in EXPORT_LEADING_COLUMNS if column in selected_columns),
        *(column for column in selected_columns if column not in EXPORT_LEADING_COLUMNS),
    ]

    headers = []
    for column in selected_columns:
        if column in EXPORT_MAIN_COLUMNS:
            headers.append(EXPORT_MAIN_COLUMNS[column])
        elif column == "quantity":
            headers.append("Остаток")
        elif column.startswith("price:"):
            headers.append(column.removeprefix("price:"))
        else:
            headers.append(warehouse_names[column.removeprefix("stock:")])

    products = (
        product_query(db, params)
        .options(
            selectinload(Product.images),
            selectinload(Product.prices),
            selectinload(Product.stocks),
            selectinload(Product.properties),
            selectinload(Product.barcodes),
        )
        .all()
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    photo_column = selected_columns.index("photo") + 1 if "photo" in selected_columns else None
    downloaded_images = download_export_images(
        [product.images[0].image_url for product in products if product.images],
        image_loader,
    ) if photo_column else {}
    if photo_column:
        worksheet.column_dimensions[get_column_letter(photo_column)].width = 16
    for product in products:
        properties: dict[str, str] = {}
        for item in product.properties:
            value = (item.value or "").strip()
            for key in (item.name, item.property_code):
                normalized_key = normalize_export_property_key(key)
                if normalized_key and value:
                    properties.setdefault(normalized_key, value)
        prices = {item.price_type: item.price_value for item in product.prices}
        stocks = {item.warehouse: item.quantity for item in product.stocks}
        main_values = {
            "code": product.code,
            "article": product.article or "",
            "photo": "",
            "name": product.name,
            "section": product.section or "",
            "product_type": product_type_names.get(product.product_type, product.product_type or ""),
            "manufacturer": product.manufacturer or "",
            "manager": product.manager or properties.get("менеджер", ""),
            "marking_code": properties.get("кодмаркировки", "") or properties.get("markingcode", ""),
            "material": product.material or "",
            "certificate": product.certificate or "",
            "barcodes": ", ".join(item.value for item in product.barcodes),
            "quantity": product.quantity,
        }
        row = []
        for column in selected_columns:
            if column in main_values:
                row.append(main_values[column])
            elif column.startswith("price:"):
                row.append(prices.get(column.removeprefix("price:"), 0))
            else:
                row.append(stocks.get(column.removeprefix("stock:"), 0))
        worksheet.append(row)
        if photo_column and product.images:
            image_content = downloaded_images.get(product.images[0].image_url)
            if image_content:
                try:
                    image = ExcelImage(BytesIO(image_content))
                    image.width = 100
                    image.height = 100
                    # Ячейка шириной 16 символов занимает около 117 px, а высотой 82,5 pt — 110 px.
                    # Небольшие смещения помещают фотографию по центру, а не у верхней левой границы.
                    image.anchor = OneCellAnchor(
                        _from=AnchorMarker(
                            col=photo_column - 1,
                            row=worksheet.max_row - 1,
                            colOff=pixels_to_EMU(8),
                            rowOff=pixels_to_EMU(5),
                        ),
                        ext=XDRPositiveSize2D(
                            cx=pixels_to_EMU(image.width),
                            cy=pixels_to_EMU(image.height),
                        ),
                    )
                    worksheet.add_image(image)
                    worksheet.row_dimensions[worksheet.max_row].height = 82.5
                except (OSError, ValueError):
                    # Повреждённое или неподдерживаемое изображение не должно прерывать весь экспорт.
                    pass

    for column_index, column in enumerate(selected_columns, start=1):
        column_letter = get_column_letter(column_index)
        if column == "photo":
            continue
        if column in EXPORT_FIXED_WIDTHS:
            width = EXPORT_FIXED_WIDTHS[column]
        elif column in EXPORT_VALUE_WIDTH_COLUMNS:
            value_lengths = (len(str(cell.value or "")) for cell in worksheet[column_letter][1:])
            width = max(len(headers[column_index - 1]), *value_lengths) + 2
        else:
            width = len(headers[column_index - 1]) + 2
        worksheet.column_dimensions[column_letter].width = width

        if column in EXPORT_FIXED_WIDTHS:
            for cell in worksheet[column_letter]:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                cell.alignment = alignment

    # Вертикальное выравнивание применяется ко всей таблице, включая заголовки,
    # обычные значения и ячейки с переносом строк.
    for row in worksheet.iter_rows():
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.vertical = "center"
            cell.alignment = alignment
    return workbook
